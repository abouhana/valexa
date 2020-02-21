from enum import Enum

import openpyxl
#from openpyxl import Worksheet
from openpyxl import Workbook
from os import path
from openpyxl.drawing.image import Image


CALIBRATION_SHEET_NAME = "Calibration_STD"
VALIDATION_SHEET_NAME = "Validation_STD"


class ColIndex(Enum):
    SERIES = 1
    LEVEL = 2
    CONCENTRATION = 3
    RESPONSE = 4


class XlsxHandler:
    def __init__(self, filename: str):
        self.workbook = openpyxl.load_workbook(filename=filename)
        self.__valid_xlsx()

    def __valid_xlsx(self):
        if CALIBRATION_SHEET_NAME not in self.workbook.sheetnames or \
                VALIDATION_SHEET_NAME not in self.workbook.sheetnames:
            raise ValueError("Bad format of Xlsx file. Use the right template")

    def get_calibration_data(self):
        return self.__get_data(CALIBRATION_SHEET_NAME)

    def get_validation_data(self):
        return self.__get_data(VALIDATION_SHEET_NAME)

    def __get_data(self, sheet_name):
        sheet: Worksheet = self.workbook[sheet_name]
        data = []
        for row in list(sheet.rows)[3:]:
            row_values = (
                row[ColIndex.SERIES.value].value,
                row[ColIndex.LEVEL.value].value,
                row[ColIndex.CONCENTRATION.value].value,
                row[ColIndex.RESPONSE.value].value
            )
            data.append(row_values)

        return data

class XlsxWriter:

    def __init__(self, profile_to_use):
        self.profile = profile_to_use

    def write_profile(self, file_name):
        workbook_file_name = file_name + '.xlsx'
        if not self.__valid_if_file_exist(workbook_file_name):
            working_workbook = Workbook()
            working_workbook.save(workbook_file_name)

        working_workbook = openpyxl.load_workbook(workbook_file_name)

        if not self.profile.model.name  in working_workbook.sheetnames:
            working_workbook.create_sheet(self.profile.model.name)

        working_sheet = working_workbook[self.profile.model.name]
        working_sheet['A1'] = self.profile.name_of_file
        working_sheet['A2'] = 'Limit of Quantification'
        working_sheet['B2'] = self.profile.min_lq
        working_sheet.add_image(Image('Linear fig.png'), 'A3')

        working_workbook.save(workbook_file_name)



    def __valid_if_file_exist(self, file_name):
        return path.exists(file_name)